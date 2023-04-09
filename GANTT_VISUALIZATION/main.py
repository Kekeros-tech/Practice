import plotly.figure_factory as ff
import socket
import pandas as pd
import os
from openpyxl.utils import get_column_letter
from _thread import start_new_thread

if not os.path.exists("images"):
    os.mkdir("images")

if not os.path.exists("dock"):
    os.mkdir("dock")

def formRes(data):
    operations = []
    distribution = []
    name = str(data[0])
    sizes = [0, 0, 0]
    sizes_full = []
    for i in range(1, len(data)):
        bufferInfo = data[i].split(';')
        resources = bufferInfo[1].split(",")
        working_hours = bufferInfo[2].split(",")
        count = 1
        for j in range(len(resources)):
            times = working_hours[j].split("--->")
            info = [bufferInfo[0], pd.to_datetime(times[0]), pd.to_datetime(times[1])]
            sizes = [max(len(str(info[k])), sizes[k]) for k in range(len(info))]
            info_full = [bufferInfo[0], resources[j], pd.to_datetime(times[0]), pd.to_datetime(times[1]), count]
            if info not in operations:
                operations.append(info)
                distribution.append(info_full)
            else:
                index = operations.index(info)
                count += 1
                distribution[index][4] = count
            #operations.append([bufferInfo[0], resources[i], times[0], times[1]])
    df = pd.DataFrame(operations,
                      columns=['Task', 'Start', 'Finish'])
    df_full = pd.DataFrame(distribution,
                           columns=['Task', 'Resources', 'Start', 'Finish', 'Count'])

    # df['duration'] = df['Start'] - df['Finish'] + 1
    df_full['Time to start'] = (df['Start'] - df['Start'].min()).dt.days
    excel_file_path = "dock/"+name+".xlsx"
    sizes_full = [len(str(x)) for x in df_full.iloc [0]]

    # for k in range(len(sizes_full)):
    #     sizes_full = [print(data[k]) for col, data in df_full.items()]
    with pd.ExcelWriter(excel_file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name='Time distribution')
        df_full.to_excel(writer, sheet_name='Full distribution')
        workbook = writer.book
        worksheet = writer.sheets['Time distribution']
        for i in range(2, worksheet.max_column + 1):
            # преобразовываем индекс столбца в его букву
            letter = get_column_letter(i)
            # получаем ширину столбца
            worksheet.column_dimensions[letter].width = sizes[i - 2] * 1.2
    # Create a figure with Plotly colorscale
    fig = ff.create_gantt(df,
                          colors='rgb(65,105,225)',
                          title="Distribution of operations",
                          showgrid_x=True,
                          showgrid_y=True,
                          height=45*len(data),
                          group_tasks=True)
    fig.write_image("images/"+name+".jpeg", width=1000, height=500)
    fig.show()
    return 0

def threaded(c):
    operations = ""
    while True:
        dataForPath = c.recv(1024)
        path_to_file = str(dataForPath)
        operations += dataForPath.decode()
        if "End" in path_to_file:
            break
        if not path_to_file:
            break
    formRes(operations.removesuffix("\nEnd").split("\n"))
    answer = bytes("End\n", "ascii")
    c.send(answer)
    c.close()

if __name__ == '__main__':
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    s.bind(('localhost', 5050))
    s.listen(5)
    while True:
        sock, addr = s.accept()
        print('Connected to :', addr[0], ':', addr[1])
        start_new_thread(threaded, (sock,))